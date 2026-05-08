"""
Gera todos os artefatos da entrega final automaticamente:
- 8 PNGs dos gráficos da D2 (notebooks/figs/)
- CSV/Excel da Matriz RACI (D1)
- CSV/Excel do Orçamento (D1)
- CSV da Matriz GUT (D3)
- Mock de board Kanban (PNG) — substitui o print do Trello

Uso:
    python scripts/gerar_artefatos.py
"""
from __future__ import annotations

from pathlib import Path
import csv

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
FIGS_DIR = ROOT / "notebooks" / "figs"
EXPORTS_D1 = ROOT / "docs" / "d1_gestao" / "exports"
EXPORTS_D3 = ROOT / "docs" / "exports"
FIGS_DIR.mkdir(parents=True, exist_ok=True)
EXPORTS_D1.mkdir(parents=True, exist_ok=True)
EXPORTS_D3.mkdir(parents=True, exist_ok=True)

PALETTE = ["#dc2626", "#f59e0b", "#16a34a", "#3b82f6", "#9333ea"]
sns.set_theme(style="darkgrid", palette=PALETTE)
plt.rcParams["figure.dpi"] = 110


# ---------------------------------------------------------------------------
# 1. PNGs dos 8 gráficos da Disciplina 2
# ---------------------------------------------------------------------------

def gerar_pngs_d2() -> None:
    csv_path = ROOT / "notebooks" / "agendamentos_tratado.csv"
    if not csv_path.exists():
        print("[D2] aviso: agendamentos_tratado.csv não existe; rode o notebook 01 antes.")
        return

    df = pd.read_csv(csv_path, sep=";", parse_dates=["data"])
    print(f"[D2] dataset carregado: {len(df)} linhas")

    def salvar(fig, nome):
        path = FIGS_DIR / f"{nome}.png"
        fig.savefig(path, bbox_inches="tight", dpi=140)
        plt.close(fig)
        print(f"  salvo: {path.relative_to(ROOT)}")

    # 1. Barras
    fig, ax = plt.subplots(figsize=(8, 5))
    df.groupby("barbeiro")["receita"].sum().sort_values(ascending=False).plot(
        kind="bar", color=PALETTE, ax=ax
    )
    ax.set_title("Receita realizada por barbeiro (R$)")
    ax.set_ylabel("R$")
    ax.set_xlabel("")
    plt.xticks(rotation=0)
    salvar(fig, "01_barras_receita_barbeiro")

    # 2. Linhas
    fig, ax = plt.subplots(figsize=(10, 4.5))
    df.groupby(df["data"].dt.date).size().plot(ax=ax, color=PALETTE[0])
    ax.set_title("Evolução diária do volume de agendamentos")
    ax.set_ylabel("Agendamentos")
    ax.set_xlabel("Data")
    salvar(fig, "02_linhas_volume_diario")

    # 3. Dispersão
    fig, ax = plt.subplots(figsize=(7, 5))
    sns.scatterplot(data=df, x="duracao_min", y="preco", hue="servico", s=80, ax=ax)
    ax.set_title("Dispersão: Duração (min) × Preço (R$)")
    salvar(fig, "03_dispersao_duracao_preco")

    # 4. Histograma
    fig, ax = plt.subplots(figsize=(7, 4.5))
    sns.histplot(df["preco"], bins=15, kde=True, color=PALETTE[0], ax=ax)
    ax.set_title("Distribuição de preços dos serviços")
    ax.set_xlabel("Preço (R$)")
    salvar(fig, "04_histograma_precos")

    # 5. Boxplot
    fig, ax = plt.subplots(figsize=(8, 5))
    sns.boxplot(data=df, x="servico", y="preco", palette=PALETTE, ax=ax)
    ax.set_title("Boxplot: Preço por serviço (outliers em destaque)")
    plt.xticks(rotation=15)
    salvar(fig, "05_boxplot_precos_servico")

    # 6. Heatmap
    numericas = df[["preco", "duracao_min", "receita", "foi_cancelado", "mes", "dia_semana"]]
    corr = numericas.corr(method="spearman")
    fig, ax = plt.subplots(figsize=(7, 6))
    sns.heatmap(corr, annot=True, fmt=".2f", cmap="RdYlGn", center=0, ax=ax)
    ax.set_title("Mapa de calor — correlação de Spearman")
    salvar(fig, "06_heatmap_correlacoes")

    # 7. Donut
    composicao = df["status"].value_counts()
    fig, ax = plt.subplots(figsize=(7, 6))
    wedges, texts, autotexts = ax.pie(
        composicao.values,
        labels=composicao.index,
        colors=PALETTE[: len(composicao)],
        autopct="%1.1f%%",
        startangle=90,
        pctdistance=0.78,
        wedgeprops=dict(width=0.4, edgecolor="white"),
    )
    for t in autotexts:
        t.set_color("white")
        t.set_fontweight("bold")
    ax.set_title("Composição dos agendamentos por status")
    salvar(fig, "07_donut_status")

    # 8. Pareto
    pareto = (
        df.groupby("servico")["receita"].sum().sort_values(ascending=False).to_frame()
    )
    pareto["acumulado_pct"] = pareto["receita"].cumsum() / pareto["receita"].sum() * 100
    fig, ax1 = plt.subplots(figsize=(9, 5))
    ax1.bar(pareto.index, pareto["receita"], color=PALETTE[0], alpha=0.85)
    ax1.set_ylabel("Receita (R$)", color=PALETTE[0])
    ax1.tick_params(axis="y", labelcolor=PALETTE[0])
    ax1.set_xlabel("Serviço")
    plt.setp(ax1.get_xticklabels(), rotation=15)
    ax2 = ax1.twinx()
    ax2.plot(pareto.index, pareto["acumulado_pct"], color=PALETTE[3], marker="o", linewidth=2)
    ax2.axhline(80, color=PALETTE[2], linestyle="--", linewidth=1.2, alpha=0.7)
    ax2.set_ylabel("% acumulado da receita", color=PALETTE[3])
    ax2.set_ylim(0, 110)
    ax2.tick_params(axis="y", labelcolor=PALETTE[3])
    plt.title("Pareto (80/20) — Receita por serviço")
    salvar(fig, "08_pareto_receita_servico")


# ---------------------------------------------------------------------------
# 2. Matriz RACI (CSV + XLSX colorido)
# ---------------------------------------------------------------------------

RACI_HEADERS = ["Pacote da EAP", "JF (PM)", "CR (Dados)", "DL (Segurança)", "GP (Pesq.Op.)", "Sponsor", "Professor"]
RACI_DATA = [
    ["1.1.1 TAP",                      "R/A", "C",  "C",  "C",  "A",  "I"],
    ["1.1.2 EAP",                      "R/A", "C",  "C",  "C",  "I",  "I"],
    ["1.1.3 Backlog (User Stories)",   "R/A", "C",  "C",  "C",  "C",  "I"],
    ["1.1.4 Matriz RACI",              "R/A", "C",  "C",  "C",  "I",  "I"],
    ["1.1.5 Cronograma Gantt",         "R/A", "C",  "C",  "C",  "C",  "I"],
    ["1.1.6 Custos",                   "R/A", "I",  "I",  "I",  "A",  "I"],
    ["1.2.1 Ata de Kickoff",           "R/A", "C",  "C",  "C",  "A",  "I"],
    ["1.2.2 Atas semanais",            "R/A", "I",  "I",  "I",  "I",  "I"],
    ["2.1.1 Modelos SQLAlchemy",       "R/A", "C",  "C",  "I",  "I",  "I"],
    ["2.1.2 Seed do banco",            "R/A", "I",  "I",  "I",  "I",  "I"],
    ["2.2.1 Rotas públicas FastAPI",   "R/A", "I",  "C",  "I",  "I",  "I"],
    ["2.2.2 Rotas administrativas",    "R/A", "I",  "C",  "I",  "C",  "I"],
    ["2.2.3 Rotas LGPD",               "C",   "I",  "R/A","I",  "I",  "I"],
    ["2.3.1 Tema dark CSS",            "R/A", "I",  "I",  "I",  "I",  "I"],
    ["2.3.2 Templates HTML",           "R/A", "I",  "I",  "I",  "I",  "I"],
    ["3.1.1 Endpoints export CSV",     "R/A", "C",  "I",  "I",  "I",  "I"],
    ["3.2.1 Notebook ETL+KPIs",        "I",   "R/A","I",  "I",  "I",  "I"],
    ["3.2.2 Notebook Visualizações",   "I",   "R/A","I",  "I",  "I",  "I"],
    ["3.3 Análise crítica D2",         "C",   "R/A","I",  "I",  "I",  "I"],
    ["4.1.1 Bcrypt + rate-limit",      "C",   "I",  "R/A","I",  "I",  "I"],
    ["4.1.2 Cabeçalhos HTTP",          "C",   "I",  "R/A","I",  "I",  "I"],
    ["4.1.3 Audit log",                "C",   "I",  "R/A","I",  "I",  "I"],
    ["4.2.1 Política de privacidade",  "C",   "I",  "R/A","I",  "C",  "I"],
    ["4.2.2 Consentimento LGPD",       "C",   "I",  "R/A","I",  "I",  "I"],
    ["4.2.3 Anonimização",             "C",   "I",  "R/A","I",  "I",  "I"],
    ["4.3 docs/d3_seguranca.md",       "I",   "I",  "R/A","I",  "I",  "I"],
    ["5.1.1 Notebook lucro (PuLP)",    "I",   "C",  "I",  "R/A","I",  "I"],
    ["5.1.2 Notebook alocação (PuLP)", "I",   "C",  "I",  "R/A","I",  "I"],
    ["5.2 What-If documentado",        "I",   "C",  "I",  "R/A","I",  "I"],
    ["6.1 Testes de aceitação",        "R",   "R",  "R",  "R",  "A",  "I"],
    ["6.2 PDF consolidado",            "R/A", "C",  "C",  "C",  "I",  "I"],
    ["6.3 Apresentação banca",         "R",   "R",  "R",  "R",  "I",  "A"],
    ["6.4 Tag v1.0 GitHub",            "R/A", "I",  "I",  "I",  "I",  "I"],
]

CORES_RACI = {
    "R":   "FF1E40AF",  # azul (Responsible)
    "A":   "FF15803D",  # verde (Accountable)
    "R/A": "FF7C3AED",  # roxo (R+A na mesma célula)
    "C":   "FFCA8A04",  # amarelo (Consulted)
    "I":   "FF6B7280",  # cinza (Informed)
}


def gerar_raci() -> None:
    csv_path = EXPORTS_D1 / "matriz_raci.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(RACI_HEADERS)
        w.writerows(RACI_DATA)
    print(f"[RACI] CSV salvo: {csv_path.relative_to(ROOT)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz RACI"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF111827")
    for col, label in enumerate(RACI_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(RACI_DATA, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx > 1 and val in CORES_RACI:
                c.fill = PatternFill("solid", fgColor=CORES_RACI[val])
                c.font = Font(bold=True, color="FFFFFFFF")

    ws.column_dimensions[get_column_letter(1)].width = 36
    for col in range(2, len(RACI_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "B2"

    legenda_inicio = len(RACI_DATA) + 4
    ws.cell(row=legenda_inicio, column=1, value="Legenda:").font = Font(bold=True)
    legenda = [
        ("R", "Responsible — quem executa"),
        ("A", "Accountable — quem aprova (apenas 1 por linha)"),
        ("R/A", "Mesma pessoa executa e aprova"),
        ("C", "Consulted — dá input antes"),
        ("I", "Informed — informado depois"),
    ]
    for i, (sigla, desc) in enumerate(legenda, start=legenda_inicio + 1):
        c1 = ws.cell(row=i, column=1, value=sigla)
        c1.fill = PatternFill("solid", fgColor=CORES_RACI[sigla])
        c1.font = Font(bold=True, color="FFFFFFFF")
        c1.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=desc)

    xlsx_path = EXPORTS_D1 / "matriz_raci.xlsx"
    wb.save(xlsx_path)
    print(f"[RACI] XLSX salvo: {xlsx_path.relative_to(ROOT)}")


# ---------------------------------------------------------------------------
# 3. Orçamento (CSV + XLSX com totais e contingência)
# ---------------------------------------------------------------------------

CUSTOS_HEADERS = ["Categoria", "Item / Recurso", "Qtd / Tempo", "Valor Unitário (R$)", "Total (R$)"]
CUSTOS_DATA = [
    ("Equipe",        "João Falbi (PM + Backend)",          "120 h",      80.00,   9600.00),
    ("Equipe",        "Calebe Ramos (Cientista de Dados)",  "50 h",      100.00,   5000.00),
    ("Equipe",        "Diego Lima (Segurança)",             "55 h",      110.00,   6050.00),
    ("Equipe",        "Guilherme Pimenta (Pesq. Operacional)","45 h",     95.00,   4275.00),
    ("Infraestrutura","Servidor em Nuvem (AWS EC2 t3.micro)","3 meses",   60.00,    180.00),
    ("Infraestrutura","Registro de domínio .com.br",        "1 ano",     40.00,     40.00),
    ("Infraestrutura","Certificado SSL (Let's Encrypt)",    "1 ano",      0.00,      0.00),
    ("Software",      "Google Workspace (e-mail+drive)",    "3 meses",   30.00,     90.00),
    ("Software",      "Conta GitHub (Free)",                "3 meses",    0.00,      0.00),
    ("Software",      "Google Colab (Free)",                "3 meses",    0.00,      0.00),
    ("Software",      "Lucidchart (trial)",                 "1 mês",      0.00,      0.00),
    ("Operacional",   "Reuniões com Sponsor",               "8 reuniões",30.00,    240.00),
    ("Operacional",   "Impressão e encadernação",           "3 cópias",  50.00,    150.00),
]


def gerar_custos() -> None:
    subtotal = sum(row[4] for row in CUSTOS_DATA)
    contingencia = round(subtotal * 0.10, 2)
    total = subtotal + contingencia

    csv_path = EXPORTS_D1 / "custos.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CUSTOS_HEADERS)
        for row in CUSTOS_DATA:
            w.writerow(row)
        w.writerow(["", "", "", "Subtotal:", f"{subtotal:.2f}"])
        w.writerow(["Contingência", "Reserva 10%", "—", "—", f"{contingencia:.2f}"])
        w.writerow(["", "", "", "TOTAL DO PROJETO:", f"{total:.2f}"])
    print(f"[CUSTOS] CSV salvo: {csv_path.relative_to(ROOT)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF111827")
    for col, label in enumerate(CUSTOS_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    cor_categoria = {
        "Equipe":         "FFDBEAFE",
        "Infraestrutura": "FFFEF3C7",
        "Software":       "FFE9D5FF",
        "Operacional":    "FFD1FAE5",
    }
    last = 1
    for r_idx, row in enumerate(CUSTOS_DATA, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if c_idx in (4, 5):
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = PatternFill("solid", fgColor=cor_categoria.get(row[0], "FFFFFFFF"))
        last = r_idx

    sub_row = last + 2
    ws.cell(row=sub_row, column=4, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=sub_row, column=5, value=subtotal).number_format = '"R$" #,##0.00'

    cont_row = sub_row + 1
    ws.cell(row=cont_row, column=1, value="Contingência").fill = PatternFill("solid", fgColor="FFFEF3C7")
    ws.cell(row=cont_row, column=2, value="Reserva 10%")
    ws.cell(row=cont_row, column=4, value="10% sobre subtotal")
    ws.cell(row=cont_row, column=5, value=contingencia).number_format = '"R$" #,##0.00'

    tot_row = cont_row + 1
    for c in range(1, 6):
        ws.cell(row=tot_row, column=c).fill = PatternFill("solid", fgColor="FF15803D")
        ws.cell(row=tot_row, column=c).font = Font(bold=True, color="FFFFFFFF")
    ws.cell(row=tot_row, column=4, value="TOTAL DO PROJETO:")
    ws.cell(row=tot_row, column=5, value=total).number_format = '"R$" #,##0.00'

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16

    xlsx_path = EXPORTS_D1 / "custos.xlsx"
    wb.save(xlsx_path)
    print(f"[CUSTOS] XLSX salvo: {xlsx_path.relative_to(ROOT)}  (Total: R$ {total:,.2f})")


# ---------------------------------------------------------------------------
# 4. Matriz GUT (CSV)
# ---------------------------------------------------------------------------

GUT_DATA = [
    (1,  "Vazamento de banco de dados",            "OWASP A02", 5, 5, 5),
    (2,  "Brute force no login admin",             "OWASP A07", 5, 5, 4),
    (3,  "SQL Injection",                          "OWASP A03", 5, 4, 5),
    (4,  "XSS armazenado",                         "OWASP A03", 5, 4, 4),
    (5,  "Session hijacking",                      "OWASP A07", 5, 4, 4),
    (6,  "Clickjacking",                           "OWASP A05", 4, 4, 4),
    (7,  "Credential stuffing",                    "OWASP A07", 4, 4, 4),
    (8,  "Stack trace exposto",                    "OWASP A05", 4, 4, 4),
    (9,  "CSRF",                                   "OWASP A01", 4, 3, 4),
    (10, "IDOR",                                   "OWASP A01", 4, 4, 3),
    (11, "Mass assignment",                        "OWASP A08", 4, 4, 3),
    (12, "DoS por inundação",                      "OWASP A04", 3, 4, 3),
    (13, "Dependência vulnerável (CVE)",           "OWASP A06", 3, 4, 4),
    (14, "Path Traversal",                         "CWE-22",    4, 4, 3),
    (15, "Open Redirect",                          "OWASP A01", 3, 3, 3),
    (16, "MIME sniffing",                          "OWASP A05", 3, 3, 2),
    (17, "Privilege escalation",                   "OWASP A01", 5, 3, 3),
    (18, "Vazamento de PII em logs",               "LGPD/CWE-532", 4, 4, 3),
    (19, "Race condition no estoque",              "CWE-362",   3, 3, 3),
    (20, "SSRF",                                   "OWASP A10", 4, 3, 2),
]


def gerar_gut() -> None:
    csv_path = EXPORTS_D3 / "matriz_gut.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["#", "Ameaça", "Referência", "G", "U", "T", "Nota (G×U×T)"])
        ordenado = sorted(GUT_DATA, key=lambda r: r[3] * r[4] * r[5], reverse=True)
        for n, ameaca, ref, g, u, t in ordenado:
            w.writerow([n, ameaca, ref, g, u, t, g * u * t])
    print(f"[GUT] CSV salvo: {csv_path.relative_to(ROOT)}")


# ---------------------------------------------------------------------------
# 5. Mock visual de board Kanban (substitui o print do Trello)
# ---------------------------------------------------------------------------

def gerar_board_kanban() -> None:
    cards = {
        "TO DO": [
            ("Roadmap pós-banca: MFA/2FA", "Diego", "Baixa"),
            ("Backup criptografado AES-256", "Diego", "Baixa"),
            ("Segregação dev/stg/prod", "João",  "Baixa"),
        ],
        "DOING": [
            ("Apresentação para banca",  "Equipe", "Alta"),
            ("Revisão final do PDF",     "João",  "Alta"),
        ],
        "REVIEW": [
            ("Análise crítica D3 — pergunta GUT", "Diego", "Alta"),
            ("Análise crítica D4 — What-if",      "Guilherme", "Média"),
        ],
        "DONE": [
            ("US-01 Login admin com bcrypt",          "Diego",     "Alta"),
            ("US-02 Agendamento online",              "João",      "Alta"),
            ("US-03 Consentimento LGPD",              "Diego",     "Alta"),
            ("US-04 a US-06 CRUD agend/barbeiro/serv","João",      "Alta"),
            ("US-07 a US-09 Produtos + estoque",      "João",      "Média"),
            ("US-10 Export CSV",                      "João",      "Média"),
            ("US-11 Rate-limit lockout",              "Diego",     "Alta"),
            ("US-12 Audit log",                       "Diego",     "Média"),
            ("US-13 Anonimização LGPD",               "Diego",     "Média"),
            ("US-14 KPIs (Notebook 01)",              "Calebe",    "Média"),
            ("US-15 Maximização lucro (Notebook 03)", "Guilherme", "Baixa"),
            ("US-16 Alocação ótima (Notebook 04)",    "Guilherme", "Baixa"),
            ("US-17 Política de privacidade",         "Diego",     "Baixa"),
            ("US-18 Cabeçalhos HTTP",                 "Diego",     "Média"),
        ],
    }
    cores_col = {"TO DO": "#374151", "DOING": "#1d4ed8", "REVIEW": "#a16207", "DONE": "#15803d"}
    cor_prio  = {"Alta": "#dc2626", "Média": "#f59e0b", "Baixa": "#6b7280"}

    fig, ax = plt.subplots(figsize=(16, 9), facecolor="#0b1020")
    ax.set_facecolor("#0b1020")
    ax.set_xlim(0, 16)
    ax.set_ylim(0, 17)
    ax.axis("off")

    fig.suptitle(
        "Board do Projeto Barbearia Invictus — GitHub Projects (snapshot da Sprint 3)",
        color="white", fontsize=15, fontweight="bold", y=0.97,
    )

    for col_idx, (col_name, cards_list) in enumerate(cards.items()):
        x0 = col_idx * 4 + 0.25
        ax.add_patch(plt.Rectangle((x0, 14.4), 3.5, 1.0,
                                   color=cores_col[col_name], alpha=0.95))
        ax.text(x0 + 1.75, 14.9, f"{col_name}  ({len(cards_list)})",
                ha="center", va="center", color="white",
                fontsize=12, fontweight="bold")

        y = 14.0
        for titulo, dono, prio in cards_list:
            altura = 1.05
            y -= altura + 0.18
            ax.add_patch(plt.Rectangle((x0, y), 3.5, altura,
                                       color="#1f2937", alpha=0.95))
            ax.add_patch(plt.Rectangle((x0, y + altura - 0.10), 3.5, 0.10,
                                       color=cor_prio[prio]))
            ax.text(x0 + 0.12, y + altura - 0.30, titulo,
                    color="white", fontsize=8.2, fontweight="bold", va="top",
                    wrap=True)
            ax.text(x0 + 0.12, y + 0.18, f"@{dono}",
                    color="#9ca3af", fontsize=7.5)
            ax.text(x0 + 3.38, y + 0.18, prio,
                    color=cor_prio[prio], fontsize=7.5, ha="right", fontweight="bold")

    out = ROOT / "docs" / "d1_gestao" / "exports" / "board_kanban.png"
    fig.savefig(out, bbox_inches="tight", dpi=130, facecolor="#0b1020")
    plt.close(fig)
    print(f"[BOARD] PNG salvo: {out.relative_to(ROOT)}")


# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print("Gerando artefatos da entrega — Barbearia Invictus")
    print("=" * 60)
    gerar_pngs_d2()
    gerar_raci()
    gerar_custos()
    gerar_gut()
    gerar_board_kanban()
    print("=" * 60)
    print("OK — todos os artefatos foram gerados.")
    print("=" * 60)


if __name__ == "__main__":
    main()
